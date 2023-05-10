from django.db.models import Max
from django.shortcuts import render,redirect
from app.models import Patient,Appointment,Doctor,Booking,Patients_Chart,Appointment_Chart,Patient_Payment,Rooms
from django.http import HttpResponse,FileResponse
import io,csv
from django.contrib import messages
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,login
from django.shortcuts import render
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime
from django.db.models import Sum
from django.core.paginator import Paginator
from django.shortcuts import render
from datetime import timezone,timedelta
from django.db.models import Q
from django.shortcuts import render



def P_Venue_Pdf(request):
    buf = io.BytesIO()
    c = canvas.Canvas(buf,pagesize=letter,bottomup=0)
    textob = c.beginText()
    textob.setTextOrigin(inch, inch)
    textob.setFont("Helvetica", 14)

    patient = Patient.objects.all()
    lines = []
    for n in patient:
        lines.append(n.email)
        lines.append(n.patient_name)
        lines.append(n.gender)
        #lines.append(n.phone)
        lines.append("=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=")
    for line in lines:
        textob.textLine(line)
    c.drawText(textob)
    c.showPage()
    c.save()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='venue.pdf')

def D_Venue_Pdf(request):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    textob = c.beginText()
    textob.setTextOrigin(inch, inch)
    textob.setFont("Helvetica", 14)

    doctor = Doctor.objects.all()
    lines = []
    for n in doctor:
        lines.append(n.doctor_name)
        lines.append(n.dob)
        lines.append(n.specialization)
        lines.append(n.experience)
        lines.append(n.email)
        # lines.append(n.phone)
        lines.append("=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=")
    for line in lines:
        textob.textLine(line)
    c.drawText(textob)
    c.showPage()
    c.save()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='venue.pdf')

def A_Venue_Pdf(request):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    textob = c.beginText()
    textob.setTextOrigin(inch, inch)
    textob.setFont("Helvetica", 14)

    appointment = Appointment.objects.all()
    lines = []
    for n in appointment:
        lines.append(n.department)
        lines.append(n.doctor_name)
        lines.append(n.appointment_data)
        lines.append(n.time_slot)
        lines.append(n.problem)
        # lines.append(n.phone)
        lines.append("=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=")
    for line in lines:
        textob.textLine(line)
    c.drawText(textob)
    c.showPage()
    c.save()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='venue.pdf')

def R_Venue_Pdf(request):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    textob = c.beginText()
    textob.setTextOrigin(inch, inch)
    textob.setFont("Helvetica", 14)

    room = Booking.objects.all()
    lines = []
    for n in room:
        lines.append(n.room_type)
        lines.append(n.patient_name)
        lines.append(n.allotment_date)
        lines.append(n.discharge_date)
        lines.append(n.doctor_name)
        # lines.append(n.phone)
        lines.append("=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=")
    for line in lines:
        textob.textLine(line)
    c.drawText(textob)
    c.showPage()
    c.save()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='venue.pdf')

def PP_Venue_Pdf(request):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    textob = c.beginText()
    textob.setTextOrigin(inch, inch)
    textob.setFont("Helvetica", 14)

    room = Patient_Payment.objects.all()
    lines = []
    for n in room:
        lines.append(n.invoice_date)
        lines.append(n.patient_name)
        lines.append(n.department)
        lines.append(n.doctor_name)
        lines.append(n.admission_date)
        lines.append(n.discharge_date)
        lines.append(n.service_name)
        lines.append(n.advance_paid)
        lines.append(n.payment_type)
        lines.append(n.card_no)
        # lines.append(n.phone)
        lines.append("=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=")
    for line in lines:
        textob.textLine(line)
    c.drawText(textob)
    c.showPage()
    c.save()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='venue.pdf')

def P_Venue_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=venue.csv'

    writer = csv.writer(response)
    patient = Patient.objects.all()
    writer.writerow(['Patient Id','Patient Name','Patient Email','Patient Gender','Patient DOB','Patient Age','Phone','Address'])
    lines = []
    for n in patient:
        writer.writerow([n.patient_id,n.patient_name,n.email,n.gender,n.dob,n.age,n.phone,n.address])

    return response

def D_Venue_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=venue.csv'

    writer = csv.writer(response)
    doctor = Doctor.objects.all()
    writer.writerow(['Doctor Id','Doctor Name','Doctor Email','Doctor Gender','Doctor DOB','Doctor Age','Specialization','Experience','Phone','Address'])
    for n in doctor:
        writer.writerow([n.doctor_id,n.doctor_name,n.email,n.gender,n.dob,n.age,n.specialization,n.experience,n.phone,n.address])

    return response

def A_Venue_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=venue.csv'

    writer = csv.writer(response)
    appointment = Appointment.objects.all()
    writer.writerow(['Patient Id','Appointment Id','Department','Doctor Name','Appointment Date','Time Slot','Token No','Problem'])
    lines = []
    for n in appointment:
        writer.writerow([n.patient_id,n.appointment_id,n.department,n.doctor_name,n.appointment_data,n.time_slot,n.token_no,n.problem])

    return response

def R_Venue_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=venue.csv'

    writer = csv.writer(response)
    room = Booking.objects.all()
    writer.writerow(['Room No','Room Type','Patient Name','Allotment Date','Discharge Date','Doctor Name'])
    lines = []
    for n in room:
        writer.writerow([n.room_no,n.room_type,n.patient_name,n.allotment_date,n.discharge_date,n.doctor_name])

    return response

def PP_Venue_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=venue.csv'

    writer = csv.writer(response)
    room = Patient_Payment.objects.all()
    writer.writerow(['Invoice id','Invoice date','Patient id','Patient name','Department','Doctor name','Admission date','Discharge date','Service name','Cost of treatment','Discount','Advance paid','Payment type','Card no'])
    lines = []
    for n in room:
        writer.writerow([n.invoice_id,n.invoice_date,n.patient_id,n.patient_name,n.department,n.doctor_name,n.admission_date,n.discharge_date,n.service_name,n.cost_of_treatment,n.discount,n.advance_paid,n.payment_type,n.card_no])

    return response

def BASE(request):
    return render(request,'base.html')

def ADD_PATIENT(request):
    if request.session.has_key('id'):
        if request.method == "POST":
            id = 1001 if Patient.objects.count() == 0 else Patient.objects.aggregate(max=Max('patient_id'))["max"] + 1
            patient_name = request.POST.get('patient_name')
            dob = request.POST.get('dob')
            age = request.POST.get('age')
            phone = request.POST.get('phone')
            email = request.POST.get('email')
            gender = request.POST.get('gender')
            address = request.POST.get('address')

            patient=Patient(
                patient_id = id,
                patient_name = patient_name,
                dob = dob,
                age = age,
                phone = phone,
                email = email,
                gender = gender,
                address = address
            )
            patient.save()
        return render(request,'patients/add_patient.html')
    else:
        return redirect('')

def ABOUT_PATIENT(request):
    if request.session.has_key('id'):
        data = Patient.objects.all()
        data1 = Patient_Payment.objects.all()
        data2 = Appointment.objects.all()
        if request.method == "POST":
            patient_name = request.POST.get('patient_name')
            dob = request.POST.get('dob')
            age = request.POST.get('age')
            phone = request.POST.get('phone')
            email = request.POST.get('email')
            gender = request.POST.get('gender')
            address = request.POST.get('address')
            invoice_id = request.POST.get('invoice_id')
            invoice_date = datetime.now()
            patient_name = request.POST.get('patient_name')
            doctor_name = request.POST.get('doctor_name')
            service_name = request.POST.get('service_name')
            cost_of_treatment = request.POST.get('cost_of_treatment')
            discount = request.POST.get('discount')
            data = Patient(patient_name,dob,age,phone,email,gender,address)
        
            data1.save()
            data.save()
            data2.save()
        paginator = Paginator(data, 3)  
        page_number = request.GET.get('page', 1)
        page = paginator.get_page(page_number)
          
        return render(request,'patients/about_patient.html',{'details':page, 'details1': data1,'details2':data2})
    else:
        return redirect('')

def ALL_PATIENT(request):
    if request.session.has_key('id'):
        data = Patient.objects.all().order_by('-patient_id')
        
        if request.method == "POST":
            patient_id = request.POST.get('patient_id')
            patient_name = request.POST.get('patient_name')
            dob = request.POST.get('dob')
            age = request.POST.get('age')
            phone = request.POST.get('phone')
            email = request.POST.get('email')
            gender = request.POST.get('gender')
            address = request.POST.get('address')

            data = Patient(patient_id,patient_name, dob, age, phone, email, gender, address)
            data.save()
        paginator = Paginator(data, 10)  # Display 10 items per page
        page_number = request.GET.get('page', 1)
        page = paginator.get_page(page_number)
        return render(request, 'patients/all_patient.html', {"details": page})
    else:
        return redirect('')


def Search_PATIENT(request):
    if request.session.has_key('id'):
        query = request.GET.get('q')
        if query:
            data = Patient.objects.all().filter(Q(patient_id__icontains=query) | Q(patient_name__icontains=query) | Q(phone__icontains=query)).order_by('-patient_id')
        else:
            data = Patient.objects.all().order_by('-patient_id')
            
        paginator = Paginator(data, 10)  # Display 10 items per page
        page_number = request.GET.get('page', 1)
        page = paginator.get_page(page_number)
        return render(request, 'patients/search.html', {"details": page, "query": query})
    else:
        return redirect('')


def EDIT_PATIENT(request,patient_id):
    patient = Patient.objects.get(patient_id=patient_id)

    return render(request,'patients/edit_patient.html',{'Patient':patient})

def Update_Patient(request,patient_id):
    patient_name = request.POST.get('patient_name')
    dob = request.POST.get('dob')
    age = request.POST.get('age')
    phone = request.POST.get('phone')
    email = request.POST.get('email')
    gender = request.POST.get('gender')
    address = request.POST.get('address')
    my_user = Patient.objects.get(patient_id=patient_id)
    my_user.patient_name = patient_name
    my_user.dob = dob
    my_user.age = age
    my_user.phone = phone
    my_user.email = email
    my_user.gender = gender
    my_user.address = address

    my_user.save()
    return redirect('all_patient')

def ADD_DOCTOR(request):
    if request.session.has_key('id'):
        if request.method == "POST":
            doctor_id = 1001 if Doctor.objects.count() == 0 else Doctor.objects.aggregate(max=Max('doctor_id'))["max"] + 1
            doctor_name = request.POST.get('doctor_name')
            dob = request.POST.get('dob')
            specialization = request.POST.get('specialization')
            experience = request.POST.get('experience')
            age = request.POST.get('age')
            phone = request.POST.get('phone')
            email = request.POST.get('email')
            gender = request.POST.get('gender')
            doctor_details = request.POST.get('doctor_details')
            address = request.POST.get('address')

            doctor = Doctor(
                doctor_id=doctor_id,
                doctor_name = doctor_name,
                dob = dob,
                specialization = specialization,
                experience = experience,
                age = age,
                phone = phone,
                email = email,
                gender = gender,
                doctor_details = doctor_details,
                address = address
            )
            doctor.save()
        return render(request,'doctors/add_doctor.html')
    else:
        return redirect('')

def ABOUT_DOCTOR(request):
    if request.session.has_key('id'):
        data = Doctor.objects.all()
        if request.method == "POST":
            doctor_id=request.POST.get('doctor_id')
            specialization = request.POST.get('specialization')
            experience = request.POST.get('experience')
            gender = request.POST.get('gender')
            address = request.POST.get('address')
            phone = request.POST.get('phone')
            dob = request.POST.get('dob')

            data = Doctor(doctor_id,specialization,experience,gender,address,phone,dob)
            data.save()
        return render(request,'doctors/about_doctor.html',{"details":data})
    else:
        return redirect('')

def ALL_DOCTOR(request):
    if request.session.has_key('id'):
        data = Doctor.objects.all().order_by('-doctor_id')
        if request.method == "POST":
            doctor_id = request.POST.get('doctor_id')
            doctor_name = request.POST.get('doctor_name')
            experience = request.POST.get('experience')
            phone = request.POST.get('phone')
            specialization = request.POST.get('specialization')

            data = Doctor(doctor_id,doctor_name,experience,phone,specialization)
            data.save()
        paginator = Paginator(data, 10)  # Display 10 items per page
        page_number = request.GET.get('page', 1)
        page = paginator.get_page(page_number)
        return render(request,'doctors/all_doctor.html',{"details":page})
    else:
        return redirect('')

def EDIT_DOCTOR(request,doctor_id):
    doctor = Doctor.objects.get(doctor_id=doctor_id)

    return render(request,'doctors/edit_doctor.html',{'Doctor':doctor})

def Update_Doctor(request,doctor_id):

    doctor_name = request.POST.get('doctor_name')
    dob = request.POST.get('dob')
    specialization = request.POST.get('specialization')
    experience = request.POST.get('experience')
    age = request.POST.get('age')
    phone = request.POST.get('phone')
    email = request.POST.get('email')
    gender = request.POST.get('gender')
    doctor_details = request.POST.get('doctor_details')
    address = request.POST.get('address')
    my_user = Doctor.objects.get(doctor_id=doctor_id)
    my_user.doctor_name = doctor_name,
    my_user.dob = dob,
    my_user.specialization = specialization,
    my_user.experience = experience,
    my_user.age = age,
    my_user.phone = phone,
    my_user.email = email,
    my_user.gender = gender,
    my_user.doctor_details = doctor_details,
    my_user.address = address

    my_user.save()
    return redirect('all_doctor')

def ADD_APPOINTMENT(request):
    if request.session.has_key('id'):
        token_no = 1 if Appointment.objects.count() == 0 else Appointment.objects.aggregate(max=Max('token_no'))["max"] + 1
        if request.method == "POST":
            patient_id = request.POST.get('patient_id')
            department = request.POST.get('department')
            doctor_name = request.POST.get('doctor_name')
            appointment_data = request.POST.get('appointment_data')
            time_slot = request.POST.get('time_slot')
            appointment_id = 101 if Appointment.objects.count() == 0 else Appointment.objects.aggregate(max=Max('appointment_id'))["max"] + 1

            problem = request.POST.get('problem')

            appointment = Appointment(
                patient_id = patient_id,
                department = department,
                doctor_name = doctor_name,
                appointment_data = appointment_data,
                time_slot = time_slot,
                token_no = token_no,
                problem = problem,
                appointment_id = appointment_id,
            )
            appointment.save()
        return render(request,'appointments/add_appointments.html')
    else:
        return redirect('')

def ALL_APPOINTMENT(request):
    if request.session.has_key('id'):
        data = Appointment.objects.all().order_by('-appointment_id')
        if request.method == "POST":
            appointment_id = request.POST.get('appointment_id')
            patient_id = request.POST.get('patient_id')
            token_no = request.POST.get('token_no')
            doctor_name = request.POST.get('doctor_name')
            problem = request.POST.get('problem')

            data = Appointment(appointment_id,patient_id,token_no,doctor_name,problem)
            data.save()
        paginator = Paginator(data, 10)  # Display 10 items per page
        page_number = request.GET.get('page', 1)
        page = paginator.get_page(page_number)
        return render(request,'appointments/all_appointments.html', {"details":page})
    else:
        return redirect('')

def ABOUT_APPOINTMENT(request):
    if request.session.has_key('id'):
        data = Appointment.objects.all()
        if request.method == "POST":
            appointment_id = request.POST.get('appointment_id')
            patient_id = request.POST.get('patient_id')
            token_no = request.POST.get('token_no')
            doctor_name = request.POST.get('doctor_name')
            problem = request.POST.get('problem')

            data = Appointment(appointment_id, patient_id, token_no, doctor_name, problem)
            data.save()
        return render(request,'appointments/details_appointments.html', {"details":data})
    else:
        return redirect('')

def EDIT_APPOINTMENT(request,appointment_id):
    appointment = Appointment.objects.get(appointment_id=appointment_id)


    return render(request,'appointments/edit_appointments.html',{'Appointment':appointment})

def Update_Appointment(request,appointment_id):
    patient_id = request.POST.get('patient_id')
    department = request.POST.get('department')
    doctor_name = request.POST.get('doctor_name')
    appointment_data = request.POST.get('appointment_data')
    time_slot = request.POST.get('time_slot')
    problem = request.POST.get('problem')
    my_user = Appointment.objects.get(appointment_id=appointment_id)
    my_user.patient_id = patient_id
    my_user.department = department
    my_user.doctor_name = doctor_name
    my_user.appointment_data = appointment_data
    my_user.time_slot = time_slot
    my_user.problem = problem

    my_user.save()
    return redirect('all_appointment')

def ADD_PAYMENT(request):
    if request.session.has_key('id'):
        if request.method == "POST":
            invoice_id = 1 if Patient_Payment.objects.count() == 0 else Patient_Payment.objects.aggregate(max=Max('invoice_id'))["max"] + 1
            current_date = datetime.now()
            invoice_date = current_date.strftime('%Y-%m-%d')
            patient_id = request.POST.get('patient_id')
            patient_name = request.POST.get('patient_name')
            doctor_name = request.POST.get('doctor_name')
            admission_date = request.POST.get('admission_date')
            discharge_date = request.POST.get('discharge_date')
            service_name = request.POST.get('service_name')
            cost_of_treatment = request.POST.get('cost_of_treatment')
            discount = request.POST.get('discount')
            advance_paid = request.POST.get('advance_paid')
            payment_type = request.POST.get('payment_type')
            card_no = request.POST.get('card_no')

            payment = Patient_Payment(
                invoice_id = invoice_id,
                invoice_date = invoice_date,
                patient_id = patient_id,
                patient_name = patient_name,
                doctor_name = doctor_name,
                cost_of_treatment = cost_of_treatment,
                service_name = service_name,
                admission_date = admission_date,
                discharge_date = discharge_date,
                discount = discount,
                advance_paid = advance_paid,
                payment_type = payment_type,
                card_no = card_no,
            )
            payment.save()
            
        return render(request,'payment/add_payment.html')
    else:
        return redirect('')
    

    
# def ADD_SERVICE(request):
#     if request.session.has_key('id'):
#         patient_id = request.POST.get('patient_id')
#         service_name = request.POST.get('service_name')
#         cost_of_treatment = request.POST.get('cost_of_treatment')
        
#         payment_service = Service_payment(
#             patient_id = patient_id,
#             service_name = service_name,
#             cost_of_treatment = cost_of_treatment,
#         )
#         payment_service.save()
#         return render(request,'payment/add_payment.html')
#     else:
#         return redirect('')

def ALL_PAYMENT(request):
    if request.session.has_key('id'):
        data = Patient_Payment.objects.all().order_by('-invoice_id')
        if request.method == "POST":
            invoice_id = request.POST.get('invoice_id')
            invoice_date = datetime.now()
            current_date = invoice_date.strftime('%Y-%m-%d')
            patient_name = request.POST.get('patient_name')
            doctor_name = request.POST.get('doctor_name')
            service_name = request.POST.get('service_name')
            cost_of_treatment = request.POST.get('cost_of_treatment')
            discount = request.POST.get('discount')

            data = Patient_Payment(invoice_id,current_date,patient_name,doctor_name,service_name,cost_of_treatment,discount)
            data.save()
        paginator = Paginator(data, 10)  # Display 10 items per page
        page_number = request.GET.get('page', 1)
        page = paginator.get_page(page_number)
        return render(request,'payment/all_payment.html', {"details":page})
    else:
        return redirect('')

def INVOICE(request):
    if request.session.has_key('id'):
        data = Patient_Payment.objects.all()
        if request.method == "POST":
            invoice_id = request.POST.get('invoice_id')
            invoice_date = datetime.now()
            patient_name = request.POST.get('patient_name')
            doctor_name = request.POST.get('doctor_name')
            service_name = request.POST.get('service_name')
            cost_of_treatment = request.POST.get('cost_of_treatment')
            discount = request.POST.get('discount')

            data = Patient_Payment(invoice_id,invoice_date,patient_name,doctor_name,service_name,cost_of_treatment,discount)
            data.save()
        return render(request,'payment/invoice.html',{'details':data})
    else:
        return redirect('')

def Invoice_E(request,invoice_id,patient_id):
    if request.session.has_key('id'):
        patient_Payment = Patient_Payment.objects.get(invoice_id=invoice_id)
        patient = Patient.objects.get(patient_id=patient_id)
        discount = patient_Payment.discount
        cost = patient_Payment.cost_of_treatment
        advance = patient_Payment.advance_paid
        dis = (discount*cost)/100
        gst = (cost*0.9)/100
        total = (gst+cost)-advance
        details= {
            'Patient_Payment':patient_Payment,
            'patient':patient,
            'total':total
        }
        print(dis)
        return render(request,'payment/invoice.html',details)
       
    else:
        return redirect('')



def invoice(request):
    invoice = Patient_Payment.objects.get()
    patient = Patient.objects.get()
    
def ADD_ROOM_A(request):
    if request.session.has_key('id'):
        if request.method == "POST":
            room_no = request.POST.get('room_no')
            room_type = request.POST.get('room_type')
            patient_name = request.POST.get('patient_name')
            allotment_date = request.POST.get('allotment_date')
            discharge_date = request.POST.get('discharge_date')
            doctor_name = request.POST.get('doctor_name')

            room = Booking(
                room_no = room_no,
                room_type = room_type,
                patient_name = patient_name,
                allotment_date = allotment_date,
                discharge_date = discharge_date,
                doctor_name = doctor_name,
            )
            room.save()
        available_rooms = Rooms.objects.filter(available=True)
        context = {
            'available_rooms': available_rooms,
        }
        
        return render(request,'room_allotment/add_room.html', context)
    else:
        return redirect('')
    
def ALL_ROOM_A(request):
    if request.session.has_key('id'):
        data = Booking.objects.all().order_by('-room_s__roomnumber')
        if request.method == "POST":
            room_no = request.POST.get('room_no')
            room_type = request.POST.get('room_type')
            patient_name = request.POST.get('patient_name')
            allotment_date = request.POST.get('allotment_date')
            discharge_date = request.POST.get('discharge_date')
            doctor_name = request.POST.get('doctor_name')
            data = Booking(room_no,room_type,patient_name,allotment_date,discharge_date,doctor_name)
            data.save()
        paginator = Paginator(data, 10)  # Display 10 items per page
        page_number = request.GET.get('page', 1)
        page = paginator.get_page(page_number)
        return render(request,'room_allotment/all_room.html', {"details":page})
    else:
        return redirect('')

def EDIT_ROOM_A(request,room_s__roomnumber):
    room = Booking.objects.get(room_s__roomnumber = room_s__roomnumber)
    return render(request,'room_allotment/edit_room.html',{'Room':room})

def Update_Room_A(request,room_s__roomnumber):
    roomnumber = request.POST.get('roomnumber')
    room_type = request.POST.get('room_type')
    patient_name = request.POST.get('patient_name')
    checkin_date = request.POST.get('checkin_date')
    checkout_date = request.POST.get('checkout_date')
    doctor_name = request.POST.get('doctor_name')
    my_user = Booking.objects.get(room_s__roomnumber=room_s__roomnumber)
    my_user.roomnumber = roomnumber
    my_user.room_type = room_type
    my_user.patient_name = patient_name
    my_user.checkin_date = checkin_date
    my_user.checkout_date = checkout_date
    my_user.doctor_name = doctor_name

    my_user.save()
    return redirect('allocated_room')

def Horizontal(request):
    if request.session.has_key('id'):
        a_data = Appointment.objects.all()
        d_data = Doctor.objects.all()
        p_total = Patient.objects.count()
        a_total = Appointment.objects.count()
        r_data = Patient_Payment.objects.aggregate(Sum('cost_of_treatment'))['cost_of_treatment__sum']
        
        details = {
            'a_data':a_data,
            'd_data':d_data,
            'p_total':p_total,
            'a_total':a_total,
            'r_data':r_data,
        }
        return render(request, 'dashboard/horizontal.html',details)
    else:
        return redirect('')
        
def Login(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if user:
                request.session['id'] = request.POST["username"]
                return redirect('dashboard_horizontal')
        else:
            return HttpResponse("Username or Password is incorrect!!!")
    return render(request, 'other_pages/login.html')

def SIGNUP(request):
    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
       # confirm_password = request.POST.get('confirm_password')
        my_user = User.objects.create_user(
            username, email, password
        )
        my_user.save()

        return redirect('')
    return render(request, 'other_pages/signup.html')

def Patients_chart(request):
    patients = Patients_Chart.objects.all()
    context = {
        'patients': patients
    }
    return render(request,'dashboard/horizontal.html',context)

def Logout(request):
    request.session.flush()
    return redirect('')

def p_delete(request,patient_id):
    user=Patient.objects.get(patient_id=patient_id)
    user.delete()
    return redirect('all_patient')

def d_delete(request,doctor_id):
    user=Doctor.objects.get(doctor_id=doctor_id)
    user.delete()
    return redirect('all_doctor')

def a_delete(request,appointment_id):
    user=Appointment.objects.get(appointment_id=appointment_id)
    user.delete()
    return redirect('all_appointment')

def r_delete(request,room_s__roomnumber):
    user=Booking.objects.get(room_s__roomnumber=room_s__roomnumber)
    user.delete()
    return redirect('allocated_room')

class GraphData(APIView):

    def get(self,request,format=None):
        patient_chart=Patients_Chart.objects.all()
        labels = []
        DefaultData = []
        for n in patient_chart:
            labels.append(n.year)
            DefaultData.append(n.patients)
        data = {
            "labels":labels,
            "DefaultData":DefaultData,
        }
        return Response(data)
    
class GraphDataLine(APIView):

    def get(self,request,format=None):
        appointment_Chart=Appointment_Chart.objects.all()
        labels = []
        DefaultData = []
        for n in appointment_Chart:
            labels.append(n.year)
            DefaultData.append(n.appointments)
        data = {
            "labels":labels,
            "DefaultData":DefaultData,
        }
        return Response(data)

def homepage(request):
	return render(request,'dashboard/index.html')

def P_Excel(request):
    customers = Patient.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.append(['Patient Id','Patient Name','Patient Email','Patient Gender','Patient DOB','Patient Age','Phone','Address'])
    for n in customers:
        ws.append([n.patient_id,n.patient_name,n.email,n.gender,n.dob,n.age,n.phone,n.address])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_orders.xlsx'
    wb.save(response)
    return response

def D_Excel(request):
    customers = Doctor.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.append(['Doctor Id','Doctor Name','Doctor Email','Doctor Gender','Doctor DOB','Doctor Age','Specialization','Experience','Phone','Address'])
    for n in customers:
        ws.append([n.doctor_id,n.doctor_name,n.email,n.gender,n.dob,n.age,n.specialization,n.experience,n.phone,n.address])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_orders.xlsx'
    wb.save(response)
    return response

def A_Excel(request):
    customers = Appointment.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.append(['Patient Id','Appointment Id','Department','Doctor Name','Appointment Date','Time Slot','Token No','Problem'])
    for n in customers:
        ws.append([n.patient_id,n.appointment_id,n.department,n.doctor_name,n.appointment_data,n.time_slot,n.token_no,n.problem])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_orders.xlsx'
    wb.save(response)
    return response

def R_Excel(request):
    customers = Booking.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.append(['Room No','Room Type','Patient Name','Allotment Date','Discharge Date','Doctor Name'])
    for n in customers:
        ws.append([n.room_no,n.room_type,n.patient_name,n.allotment_date,n.discharge_date,n.doctor_name])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_orders.xlsx'
    wb.save(response)
    return response

def PP_Excel(request):
    customers = Patient_Payment.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.append(['Invoice id','Invoice date','Patient id','Patient name','Department','Doctor name','Admission date','Discharge date','Service name','Cost of treatment','Discount','Advance paid','Payment type','Card no'])
    for n in customers:
        ws.append([n.invoice_id,n.invoice_date,n.patient_id,n.patient_name,n.department,n.doctor_name,n.admission_date,n.discharge_date,n.service_name,n.cost_of_treatment,n.discount,n.advance_paid,n.payment_type,n.card_no])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_orders.xlsx'
    wb.save(response)
    return response

def upcoming_appointments(request):
    if request.session.has_key('id'):
            today = datetime.now(timezone.utc)
            next_week = today + timedelta(days=7)
            data = Appointment.objects.filter(
                appointment_data__range=(today, next_week)
            ).order_by('appointment_data')
            return render(request, 'appointments/appointments_notification.html', {'details':data})
    else:
        return redirect('')

def ADD_ROOM(request):
    if request.session.has_key('id'):
        
        if request.method == "POST":
            roomnumber = request.POST.get('roomnumber')
            room_type = request.POST.get('room_type')
            room = Rooms(
                roomnumber = roomnumber,
                room_type = room_type,
            )
            room.save()
            
        return render(request,'room/add_room.html')
    else:
        return redirect('')
    
def ALL_ROOM(request):
    if request.session.has_key('id'):
        data = Rooms.objects.all().order_by('-roomnumber')
        if request.method == "POST":
            roomnumber = request.POST.get('roomnumber')
            room_type = request.POST.get('room_type')
            data = Rooms(roomnumber,room_type)
            data.save()
        paginator = Paginator(data, 10)  # Display 10 items per page
        page_number = request.GET.get('page', 1)
        page = paginator.get_page(page_number)
        
        return render(request,'room/all_room.html', {"details":page})
    else:
        return redirect('')
    
    
def available_rooms(request):
    if request.session.has_key('id'):
        # Get all rooms that are not currently booked
        available_rooms = Rooms.objects.filter(~Q(roomnumber__in=Booking.objects.filter(checkout_date__gt=datetime.now()).values_list('room_s__roomnumber', flat=True)))
        
        return render(request, 'room/avalible_room.html',{'available_rooms':available_rooms})
    else:
        return redirect('')

def allocated_rooms(request):
    if request.session.has_key('id'):
        allocated_bookings = Booking.objects.filter(checkin_date__lte=datetime.now(), checkout_date__gte=datetime.now())
    
        return render(request, 'room/allocated_rooms.html', {'allocated_bookings': allocated_bookings})
    else:
        return redirect('')

    
def BookRoom(request):
    
    if request.session.has_key('id'):
        if request.method == "POST":
            roomnumber = request.POST.get('roomnumber')
            patient_name = request.POST.get('patient_name')
            checkin_date  = request.POST.get('checkin_date ')
            checkout_date = request.POST.get('checkout_date')
            doctor_name = request.POST.get('doctor_name')
            room = Rooms.objects.get(roomnumber=roomnumber)
            
            if checkin_date:
                booking = Booking(
                    room_s=room,
                    patient_name=patient_name,
                    checkin_date=checkin_date,
                    checkout_date=checkout_date,
                    doctor_name=doctor_name,
                )
                booking.save()
                room.status = 'allocated'
                room.save()
                return render(request, 'room/bookroom.html')
            else:
                # Handle the case where checkin_date is empty
                return render(request, 'room/bookroom.html', {'roomnumbers': roomnumbers,})
                # Render error page if no available rooms found
                

        else:
            # Render booking form with available room options
            available_rooms = Rooms.objects.filter(~Q(roomnumber__in=Booking.objects.filter(checkout_date__gt=datetime.now()).values_list('room_s__roomnumber', flat=True)))
            roomnumbers = [room.roomnumber for room in available_rooms]
            return render(request, 'room/bookroom.html', {'roomnumbers': roomnumbers,})
    else:
        return redirect(' ')

def Help(request):
    return render(request,'dashboard/help.html')


def Setting_page(request):
    
    return render(request,'dashboard/setting_page.html')

def faq(request):
    questions = [
        {
            'question': 'What is a Hospital Management System?',
            'answer': 'A Hospital Management System (HMS) is a software application that helps healthcare organizations manage their administrative, financial, and clinical functions.'
        },
        {
            'question': 'What are the benefits of using a Hospital Management System?',
            'answer': 'The benefits of using a Hospital Management System include improved patient care, reduced workload for healthcare professionals, enhanced data security, and streamlined administrative processes.'
        },
        {
            'question': 'How does a Hospital Management System work?',
            'answer': 'A Hospital Management System typically consists of several modules, such as patient management, appointment scheduling, billing and invoicing, and inventory management. These modules are interconnected and share data to provide a comprehensive view of hospital operations.'
        },
        {
            'question': 'What are the key features of a Hospital Management System?',
            'answer': 'The key features of a Hospital Management System include patient management, appointment scheduling, medical record management, inventory management, billing and invoicing, and reporting and analytics.'
        },
        {
            'question': 'Can Hospital Management Systems be customized?',
            'answer': 'Yes, Hospital Management Systems can be customized to suit the specific needs of a healthcare organization. Customization can include the addition or removal of modules, changes to user interfaces, and integration with third-party software.'
        },
         {
            'question': 'What are the benefits of using a Hospital Management System?',
            'answer': 'The benefits of using a Hospital Management System include improved patient care, reduced workload for healthcare professionals, enhanced data security, and streamlined administrative processes.'
        },
        # Add more questions and answers as needed
    ]
    return render(request, 'other_pages/faq.html', {'questions': questions})


def handler404(request):
    return render(request, 'other_pages/404.html', status=404)